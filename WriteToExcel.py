print("Starting pyExtract")

# Import all libraries and modules

# Standard Libraries
import os
import shutil
from datetime import datetime, timedelta
import dateutil.parser



# Libraries that we need to install

# used to read/write from excel
# Install with "pip install openpyxl"
from openpyxl import Workbook, load_workbook

# pywin32 allows us to access the outlook native application
# Install with "pip install pywin32"
import win32com.client

# used to convert outlook .msg files to pdf and extract attachments
# Install with "pip install msgtopdf"
# Uses wkhtmltopdf tool that must be installed and added to PATH
from msgtopdf import Msgtopdf

# Importing for file dialog
# Install with "pip install tk-tools"
from tkinter import Tk
from tkinter.filedialog import askdirectory

# Check if Outlook is running
# Install with "pip install psutil"
import psutil


# Note: Outlook desktop app must be installed and running
# https://www.geeksforgeeks.org/how-to-check-if-an-application-is-open-in-python/
flag = 0
for i in psutil.process_iter():
    if i.name() == "OUTLOOK.EXE":
        flag = 1
        break

if flag == 0:
    print("Outlook is not running. Start Outlook and try again")
    exit()














# Start a session with outlook
# https://www.codeforests.com/2020/06/04/python-to-read-email-from-outlook/
outlook = win32com.client.Dispatch("outlook.application")
mapi = outlook.GetNamespace("MAPI")

# Get email accounts
print("")
print("Using numbers, choose an email account to begin or type 'quit' to exit")
for idx, account in enumerate(mapi.Folders):
    #index starts from 1
    print("\t", idx+1, account)

my_account = input("Account: ")
if my_account == "quit":
    exit()
my_account = int(my_account)
my_account_name = mapi.Folders(my_account)


# Get folders
print("")
print("Using numbers, choose a folder or type 'quit' to exit")
for idx, folder in enumerate(mapi.Folders(my_account).Folders): 
    print("\t", idx+1, folder)

my_folder = input("Folder: ")
if my_folder == "quit":
    exit()
my_folder = int(my_folder)
my_folder_name = mapi.Folders(my_account).Folders(my_folder)


# Another level?
print("")
print("Using numbers, make a selection or type 'quit' to exit")
print("\t", "1 Run pyExtract on this folder")
print("\t", "2 Find another folder")

my_selection = input("My Selection: ")
if my_selection == 'quit':
    exit()
elif my_selection == "1":
    messages = mapi.Folders(my_account).Folders(my_folder).Items
    print("")
    print("You have chosen:", "\t", my_account_name, ">", my_folder_name)
elif my_selection == "2":
    my_selection == int(my_selection)
    # Another level?
    print("")
    print("Using numbers, make a selection or type 'quit' to exit")
    # finding a subfolder in : 
    for i in range(len(mapi.Folders(my_account).Folders(my_folder).Folders)):
        print("\t", i+1, mapi.Folders(my_account).Folders(my_folder).Folders[i])
        my_child = input("Folder: ")
        if my_child == "quit":
            exit()
        my_child = int(my_child)
        my_child_name = mapi.Folders(my_account).Folders(my_folder).Folders[i]
        messages = mapi.Folders(my_account).Folders(my_folder).Folders[i].Items
        print("")
        print("You have chosen:", "\t", my_account_name, ">", my_folder_name, ">", my_child_name)


print(messages)
print(type(messages))

my_confirm = input("Enter 'y' to accept or 'quit' to start over: ")
if my_confirm == "quit":
    exit()











## TODO
# Include filter for attachments
# Make it recursive

# Another level?
print("")
print("Using numbers, make a selection or type 'quit' to exit")
print("\t", "1 Run all")
print("\t", "2 'n' most recent")
print("\t", "3 By age")
print("\t", "4 Filter by Subject")
print("\t", "5 Filter by Sender")
my_filter = input("My Selection: ")
if my_filter == "quit":
    exit()
my_filter = int(my_filter)

if my_filter == 1:
    msg_count = len(list(messages))
    print("There are ", msg_count, ". Enter y to confirm:")
    my_msg_count = msg_count
elif my_filter == 2:
    my_msg_count = int(input("Select n most recent messages to extract: "))
    messages.Sort("[ReceivedTime]", Descending=True)
    msg_count = len(list(messages)[:my_msg_count])
elif my_filter == 3:
    mm, dd, hh = input("Select oldest email in the format 'month day hour' (eg 3 14 11) seperated by spaces. ").split()
    mm = int(mm)
    dd = int(dd)
    hh = int(hh)
    today = datetime.today()
    start_time = today.replace(month=mm, day=dd, hour=hh, minute=0, second=0).strftime('%Y-%m-%d %H:%M %p')
    #today 12am
    end_time = today.replace(hour=0, minute=0, second=0).strftime('%Y-%m-%d %H:%M %p')
    messages = messages.Restrict("[ReceivedTime] >= '" + start_time + "' And [ReceivedTime] <= '" + end_time + "'")
    print("Start Time: ", start_time)
    print("End Time: ", end_time)
    msg_count = len(list(messages))
    print("There are ", msg_count, ". Enter y to confirm:")
    my_msg_count = msg_count
elif my_filter == 4:
    my_subject = input("Filter by subject: ")
    my_subject = "[Subject] = '" + my_subject + "'"
    messages = messages.Restrict(my_subject)
    msg_count = len(list(messages))
    print("There are ", msg_count, ". Enter y to confirm:")
    my_msg_count = msg_count
elif my_filter == 5:
    my_sender = input("Filter by sender (email address): ")
    my_sender = "[SenderEmailAddress] = '" + my_sender + "'"
    messages = messages.Restrict(my_sender)
    msg_count = len(list(messages))
    print("There are ", msg_count, ". Enter y to confirm:")
    my_msg_count = msg_count
else:
    exit()









print("Choose a folder or press cancel")
# Ask the user to choose a location
path_parent = askdirectory(title='Select Folder') # shows dialog box and return the path
if not path_parent:
    print("No folder selected. Goodbye")
    exit()

# Creating new folder structure and saving full path names for future use
export_foldername = 'Export {:%Y-%b-%d %H-%M-%S}'.format(datetime.now())
export_folderpath = os.path.join(path_parent, export_foldername)
os.mkdir(export_folderpath)
tmp_attachments_foldername = "Attachments"
tmp_pdfs_foldername = "PDF"
tmp_msgs_foldername = "MSG"
attachments_folderpath = os.path.join(export_folderpath, tmp_attachments_foldername)
pdfs_folderpath = os.path.join(export_folderpath, tmp_pdfs_foldername)
msgs_folderpath = os.path.join(export_folderpath, tmp_msgs_foldername)
os.mkdir(attachments_folderpath)
os.mkdir(pdfs_folderpath)
os.mkdir(msgs_folderpath)






















print("Creating Excel Template")
# Creating Excel Template with column headers
wb = Workbook()
ws = wb.active

ws["A1"] = "PDF File"
ws["B1"] = "MSG File"
ws["C1"] = "Date"
ws["D1"] = "Time"
ws["E1"] = "Subject"
ws["F1"] = "Body"
ws["G1"] = "From: (Name)"
ws["H1"] = "From: (Address)"
ws["I1"] = "To: (Name)"
ws["J1"] = "To: (Address)"
ws["K1"] = "CC: (Name)"
ws["L1"] = "CC: (Address)"
ws["M1"] = "BCC: (Name)"
ws["N1"] = "BCC: (Address)"
ws["O1"] = "Importance"
ws["P1"] = "Sensitivity"
ws["Q1"] = "Attachment 1"
ws["R1"] = "Attachment 2"
ws["S1"] = "Attachment 3"
ws["T1"] = "Attachment 4"
ws["U1"] = "Attachment 5"
ws["V1"] = "Attachment 6 "
ws["W1"] = "Attachment 7"
ws["X1"] = "Attachment 8"
ws["Y1"] = "Attachment 9"
ws["Z1"] = "Attachment 10"

index_fullfilename = os.path.join(export_folderpath, "Index.xlsx")
wb.save(index_fullfilename)
wb.close()

























'''
Create new folder and structure
'''

# Opening index spreadsheet
my_wb = load_workbook(filename=index_fullfilename)
my_ws = my_wb.active

# Starting on row two, process the 10 most recent messages
icell = 2

total_messages = len(list(messages)[:my_msg_count])
current_message = 0
for msg in list(messages)[:my_msg_count]:

    # Progress bar
    current_message += 1
    print("Working. Processing message ", current_message, "out of ", total_messages)

    my_PdfFile = ""
    my_MsgFile = ""
    # ReceivedTime includes timezone so we need to clean it up and seperate date and time
    my_Date = str(msg.ReceivedTime)                     
    my_Date = dateutil.parser.parse(my_Date)
    my_Time = my_Date.strftime("%H-%M-%S")
    my_Date = str(my_Date.date())

    # Getting other info from message
    my_Sub = msg.Subject
    my_Body = msg.Body
    my_From_Name = msg.SenderName
    my_From_Address = msg.SenderEmailAddress
    my_To_Name = msg.ReceivedByName
    my_To_Address = msg.ReceivedByName   # Delete?

    # Building CC and BCC strings
    # https://learn.microsoft.com/en-us/office/vba/api/outlook.olmailrecipienttype
    # Originator = 0
    # To = 1
    # CC = 2
    # BCC = 3
    my_CC_Address = ""
    my_CC_Name = ""
    my_BCC_Address = ""
    my_BCC_Name = ""
    for person in msg.Recipients:
        if person.type == 2:
            my_CC_Address += person.Address + ", "
            my_CC_Name += person.Name + ", "
        elif person.type == 3:
            my_BCC_Address += person.Address + ", "
            my_BCC_Name += person.Name + ", "
        else:
            pass
    
    # Gettin enumerated values from importance
    # https://learn.microsoft.com/en-us/office/vba/api/outlook.olimportance
    # 0 = Low
    # 1 = Normal
    # 2 = High
    my_Importance = ""
    if msg.Importance == 0:
        my_Importance = "Low"
    elif msg.Importance == 1:
        my_Importance = "Normal"
    elif msg.Importance == 2:
        my_Importance = "High"
    else:
        pass

    # Gettin enumerated values from sensitivity
    # https://learn.microsoft.com/en-us/office/vba/api/outlook.olsensitivity
    # 0 = Normal
    # 1 = Personal
    # 2 = Private
    # 3 = Confidential
    my_Sensitivity = ""
    if msg.Sensitivity == 0:
        my_Sensitivity = "Normal"
    elif msg.Sensitivity == 1:
        my_Sensitivity = "Personal"
    elif msg.Sensitivity == 2:
        my_Sensitivity = "Private"
    elif msg.Sensitivity == 3:
        my_Sensitivity = "Confidential"
    else:
        pass

    # Placing Attachments in spreadsheet as hyperlinks
    try:
        s = msg.sender
        j = 0
        for attachment in msg.Attachments:
            attachment.SaveAsFile(os.path.join(attachments_folderpath, attachment.FileName))
            att_cell = my_ws.cell(row = icell, column = 17 + j)
            att_cell.hyperlink = os.path.join(attachments_folderpath, attachment.FileName)
            att_cell.value = attachment.FileName
            att_cell.style = "Hyperlink"
            j += 1
    except Exception as e:
        print("error when saving the attachment" + str(e))

    # assigning variables to cells
    cell1 = my_ws.cell(row = icell, column = 1)
    cell2 = my_ws.cell(row = icell, column = 2)
    cell3 = my_ws.cell(row = icell, column = 3)
    cell4 = my_ws.cell(row = icell, column = 4)
    cell5 = my_ws.cell(row = icell, column = 5)
    cell6 = my_ws.cell(row = icell, column = 6)
    cell7 = my_ws.cell(row = icell, column = 7)
    cell8 = my_ws.cell(row = icell, column = 8)
    cell9 = my_ws.cell(row = icell, column = 9)
    cell10 = my_ws.cell(row = icell, column = 10)
    cell11 = my_ws.cell(row = icell, column = 11)
    cell12 = my_ws.cell(row = icell, column = 12)
    cell13 = my_ws.cell(row = icell, column = 13)
    cell14 = my_ws.cell(row = icell, column = 14)
    cell15 = my_ws.cell(row = icell, column = 15)
    cell16 = my_ws.cell(row = icell, column = 16)

    # Placing message data into spreadsheet
    cell1.value = my_PdfFile
    cell2.value = my_MsgFile
    cell3.value = my_Date
    cell4.value = my_Time
    cell5.value = my_Sub
    cell6.value = my_Body
    cell7.value = my_From_Name
    cell8.value = my_From_Address
    cell9.value = my_To_Name
    cell10.value = my_To_Address
    cell11.value = my_CC_Name
    cell12.value = my_CC_Address
    cell13.value = my_BCC_Name
    cell14.value = my_BCC_Address
    cell15.value = my_Importance
    cell16.value = my_Sensitivity

    # Cleaning the subject line, making it shorter, and adding a date and time stamp to avoid overwriting duplicates
    # Deleting anything that isn't a number or character. Starting string with msg_ in case there are restricted characters
    # at the beginning of the string. 
    my_clean_sub = "msg_"
    for ch in my_Sub:
        if ch.isalnum():
            my_clean_sub += ch
    
    # Limiting file name to 20 characters before date and time stamp
    my_clean_sub = my_clean_sub[0:20]

    # Adding date and time stamp
    my_new_sub = my_clean_sub + "_" + my_Date + "_"+ my_Time
    msg_fullfilename = msgs_folderpath + "//" + my_new_sub + ".msg"
    msg.saveas(msg_fullfilename, 3)

    # Now we have a clean msg folder. Lets convert one at a time, move to the pdf folder then
    # delete the old. Then we can add hyperlinks to the .msg and .pdf
    email = Msgtopdf(msg_fullfilename)
    email.email2pdf()

    tmp_pdffolder = msgs_folderpath + "//" + my_new_sub
    tmp_pdffolder = os.path.join(msgs_folderpath, tmp_pdffolder)

    tmp_pdf_fullfilename = my_new_sub + ".pdf"
    pdf_fullfilename = os.path.join(tmp_pdffolder, tmp_pdf_fullfilename)

    shutil.copy(pdf_fullfilename, pdfs_folderpath)
    newpdf_fullfilename = os.path.join(pdfs_folderpath, tmp_pdf_fullfilename)

    # Delete old pdf folders for cleanup
    shutil.rmtree(tmp_pdffolder)

    # Adding hyperlinks
    cell1.hyperlink = newpdf_fullfilename
    cell1.value = "PDF"
    cell1.style = "Hyperlink"

    cell2.hyperlink = msg_fullfilename
    cell2.value = "MSG"
    cell2.style = "Hyperlink"

    # Next message
    icell += 1

# Save and close spreadsheet
my_wb.save(index_fullfilename)
my_wb.close()
print("Success. Goodbye")
